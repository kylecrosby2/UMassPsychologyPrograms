import random
import decimal
import openpyxl as xl

workbook = xl.load_workbook('ExperimentFile1.xlsx')
sheet = workbook.active
m_row = sheet.max_row
cell_list = []
for i in range(1, m_row + 1):
    cell = sheet.cell(row=i, column=1)
    cell_list.append(cell.value)
print(cell_list)
while True:
    rand_time = float(decimal.Decimal(random.randrange(950, 1150)) / 1000)
    repeat_num = False
    for i in cell_list:
        if i == rand_time:
            repeat_num = True
    if repeat_num is False:
        sheet['A%s' % (m_row + 1)] = rand_time
        break
print(rand_time)
workbook.save('ExperimentFile1.xlsx')
